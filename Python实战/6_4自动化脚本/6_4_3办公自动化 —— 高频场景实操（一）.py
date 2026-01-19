# import os
# import openpyxl
#
#
# def batch_read_excel(folder_path):
#     if not os.path.exists(folder_path):
#         print(f"错误：文件夹「{folder_path}」不存在！")
#         return
#
#     excel_file_list = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
#
#     if len(excel_file_list) == 0:
#         print(f"提示：文件夹「{folder_path}」内无.xlsx文件！")
#         return
#
#     for file_name in excel_file_list:
#         excel_file_path = os.path.join(folder_path, file_name)
#         print(f"\n========== 开始读取文件：{file_name} ==========")
#         try:
#             wb = openpyxl.load_workbook(excel_file_path, read_only=True)
#             ws = wb.active
#
#             max_row = ws.max_row
#             max_column = ws.max_column
#
#             print(f"工作表名：{ws.title}")
#             print(f"数据范围：第1行 - 第{max_row}行，第1列 - 第{max_column}列")
#             print("前5行数据（表头+内容）：")
#
#             for row in range(1, min(max_row + 1, 6)):
#                 row_data = []
#                 for col in range(1, max_column + 1):
#                     cell_value = ws.cell(row=row, column=col).value
#                     row_data.append(cell_value if cell_value is not None else "")
#                 print(f"第{row}行：{row_data}")
#
#             wb.close()
#             print(f"========== 文件「{file_name}」读取完成 ==========\n")
#         except Exception as e:
#             print(f"错误：读取文件「{file_name}」失败，错误信息：{e}")
#             continue
#
# if __name__ == '__main__':
#     target_folder = "D:\\Projects\\AiStudy\\Python实战\\6_4自动化脚本\\办公文件\\Excel测试文件夹"
#     batch_read_excel(target_folder)
import os.path


# import os
# import openpyxl
#
#
# def batch_create_excel(folder_path, file_count, data_list):
#     os.makedirs(folder_path, exist_ok=True)
#
#     for i in range(1, file_count + 1):
#         excel_file_name = f"办公报表_{i}.xlsx"
#         excel_file_path = os.path.join(folder_path, excel_file_name)
#         try:
#             wb = openpyxl.Workbook()
#             ws = wb.active
#
#             ws.title = "核心数据"
#             for row_index, row_data in enumerate(data_list, start=1):
#                 for col_index, cell_data in enumerate(row_data, start=1):
#                     ws.cell(row=row_index, column=col_index, value=cell_data)
#
#             wb.save(excel_file_path)
#             wb.close()
#             print(f"成功创建Excel文件：{excel_file_name}")
#         except Exception as e:
#             print(f"错误：创建文件「{excel_file_name}」失败，错误信息：{e}")
#             continue
#
#
# if __name__ == '__main__':
#     target_folder = "D:\\Projects\\AiStudy\\Python实战\\6_4自动化脚本\\办公文件\\批量生成Excel"
#     create_count = 3
#
#     common_data = [["姓名", "部门", "入职日期", "薪资"],
#                    ["张三", "技术部", "2024-01-01", 8000],
#                    ["李四", "市场部", "2024-02-01", 6000]]
#     batch_create_excel(target_folder, create_count, common_data)

# import os
# import shutil
#
# def auto_organize_folder(source_folder):
#     if not os.path.isdir(source_folder):
#         print(f"错误：文件夹「{source_folder}」不存在！")
#         return
#
#     file_type_map = {
#         # 文档类
#         ".txt": "文本文档",
#         ".docx": "Word文档",
#         ".doc": "Word文档",
#         ".pdf": "PDF文档",
#         ".ppt": "PPT文档",
#         ".pptx": "PPT文档",
#         # 表格类
#         ".xlsx": "Excel文档",
#         ".xls": "Excel文档",
#         # 图片类
#         ".jpg": "图片文件",
#         ".jpeg": "图片文件",
#         ".png": "图片文件",
#         ".gif": "图片文件",
#         # 压缩包类
#         ".zip": "压缩包文件",
#         ".rar": "压缩包文件",
#         # 其他类
#         ".exe": "可执行文件",
#         ".csv": "CSV文件"
#     }
#
#     for item_name in os.listdir(source_folder):
#         item_full_path = os.path.join(source_folder, item_name)
#         if os.path.isdir(item_full_path):
#             print(f"跳过子目录：{item_name}")
#             continue
#
#         if item_name.startswith("."):
#             continue
#
#         file_suffix = os.path.splitext(item_name)[1].lower()
#
#         if file_suffix in file_type_map:
#             target_folder_name = file_type_map[file_suffix]
#         else:
#             target_folder_name = "其他文件"
#
#         target_folder_path = os.path.join(source_folder, target_folder_name)
#
#         os.makedirs(target_folder_path, exist_ok=True)
#
#         target_file_path = os.path.join(target_folder_path, item_name)
#
#         try:
#             shutil.move(item_full_path, target_file_path)
#             print(f"成功移动：{item_name} → {target_folder_name}/{item_name}")
#         except Exception as e:
#             print(f"错误：移动文件「{item_name}」失败，错误信息：{e}")
#             continue
#
#     print("\n========== 文件夹自动整理完成！ ==========")
#
# if __name__ == '__main__':
#     messy_folder = "D:\\Projects\\AiStudy\\Python实战\\6_4自动化脚本\\办公文件\\杂乱文件夹"
#     auto_organize_folder(messy_folder)
import os
import openpyxl
import shutil


def auto_organize_folder(source_folder):
    file_type_map = {
        ".xlsx": "Excel文档",
        ".xls": "Excel文档",
        ".txt": "文本文档",
        ".docx": "Word文档",
        ".pdf": "PDF文档",
        ".jpg": "图片文件",
        ".zip": "压缩包文件"
    }
    if not os.path.isdir(source_folder):
        print(f"错误：文件夹「{source_folder}」不存在！")
        return False

    for item_name in os.listdir(source_folder):
        item_full_path = os.path.join(source_folder, item_name)
        if os.path.isdir(item_full_path):
            continue
        file_suffix = os.path.splitext(item_name)[1].lower()
        target_folder_name = file_type_map.get(file_suffix, "其他文件")
        target_folder_path = os.path.join(source_folder, target_folder_name)
        os.makedirs(target_folder_path, exist_ok=True)
        target_file_path = os.path.join(target_folder_path, item_name)

        try:
            shutil.move(item_full_path, target_file_path)
        except:
            continue
    print("文件夹整理完成！")
    return True


def batch_summary_excel(folder_path):
    summary_data = []
    excel_folder = os.path.join(folder_path, "Excel文档")
    if not os.path.isdir(excel_folder):
        print("无Excel文档文件夹！")
        return summary_data

    excel_file_list = [f for f in os.listdir(excel_folder) if f.endswith(".xlsx")]
    for file_name in excel_file_list:
        excel_path = os.path.join(excel_folder, file_name)
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            max_row = ws.max_row
            max_col = ws.max_column
            # 跳过表头，读取数据行（按需调整）
            for row in range(2, max_row + 1):
                row_data = [file_name]  # 先添加文件名，便于溯源
                for col in range(1, max_col + 1):
                    cell_val = ws.cell(row=row, column=col).value or ""
                    row_data.append(cell_val)
                summary_data.append(row_data)
            wb.close()
        except:
            continue
    return summary_data


if __name__ == '__main__':
    target_folder = "D:\\Projects\\AiStudy\\Python实战\\6_4自动化脚本\\办公文件\\综合测试文件夹"
    organize_success = auto_organize_folder(target_folder)
    if not organize_success:
        exit()

    excel_summary = batch_summary_excel(target_folder)
    print(f"\nExcel数据汇总完成，共汇总 {len(excel_summary)} 行数据")
    print("前3行汇总数据：")
    for row in excel_summary[:3]:
        print(row)